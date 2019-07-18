import sys
from pandas.compat import reload
reload(sys)
from bs4 import BeautifulSoup
import re
import urllib
from openpyxl import Workbook

#得到页面全部内容
def askURL(url):
  request = urllib.request.Request(url)#发送请求
  try:
    response = urllib.request.urlopen(request)#取得响应
    html= response.read()#获取网页内容
  except (urllib.request.HTTPError, urllib.request.URLError) as e:
    if hasattr(e,"code"):
      print
      e.code
    if hasattr(e,"reason"):
      print
      e.reason
  return html

#获取相关内容
def getData(baseurl):
  findLink=re.compile(r'<a href="(.*?)">')#找到影片详情链接
  findImgSrc=re.compile(r'<img.*src="(.*jpg)"',re.S)#找到影片图片
  findTitle=re.compile(r'<span class="title">(.*)</span>')#找到片名
  #找到评分
  findRating=re.compile(r'<span class="rating_num" property="v:average">(.*)</span>')
  #找到评价人数
  findJudge=re.compile(r'<span>(\d*)人评价</span>')
  #找到概况
  findInq=re.compile(r'<span class="inq">(.*)</span>')
  #找到影片相关内容：导演，主演，年份，地区，类别
  findBd=re.compile(r'<p class="">(.*?)</p>',re.S)
  #去掉无关内容
  remove=re.compile(r'              |\n|</br>|\.*')
  datalist=[]
  for i in range(0,10):
    url=baseurl+str(i*25)
    html=askURL(url)
    soup = BeautifulSoup(html, "html.parser")
    for item in soup.find_all('div',class_='item'):#找到每一个影片项
      data=[]
      item=str(item)#转换成字符串
      link=re.findall(findLink,item)[0]
      data.append(link)#添加详情链接
      imgSrc=re.findall(findImgSrc,item)[0]
      data.append(imgSrc)#添加图片链接
      titles=re.findall(findTitle,item)
      #片名可能只有一个中文名，没有外国名
      if(len(titles)==2):
        ctitle=titles[0]
        data.append(ctitle)#添加中文片名
        otitle=titles[1].replace(" / ","")#去掉无关符号
        otitle=titles[1].replace("/","")
        data.append(otitle)#添加外国片名
      else:
        data.append(titles[0])#添加中文片名
        data.append(' ')#留空
      rating=re.findall(findRating,item)[0]
      data.append(rating)#添加评分
      judgeNum=re.findall(findJudge,item)[0]
      data.append(judgeNum)#添加评论人数
      inq=re.findall(findInq,item)
      #可能没有概况,添加概况
      if len(inq)!=0:
        inq=inq[0].replace("。","")#去掉句号
        data.append(inq)#添加概况
      else:
        data.append(' ')#留空
      #添加导演
      bd=re.findall(findBd,item)[0]
      bd=re.sub(remove,"",bd)
      bd=re.sub('<br>'," ",bd)#去掉<br>
      bd=re.sub('/'," ",bd)#替换/
      words=bd.split("<br >")
      directions=words[0].split("   ")
      #写入导演和主演
      if(directions.__len__()<2):
        data.append(directions[0].replace("导演: ",""))
        data.append("未显示主演")
      else:
        data.append(directions[0].replace("导演: ",""))
        data.append(directions[1].replace("主演: ",""))
      ancitions = words[1].split("   ")
      data.append(ancitions[0])#添加年份
      data.append(ancitions[1])#添加地区
      data.append(ancitions[2])#添加类别
      datalist.append(data)
  return datalist

#将相关数据写入excel中
def saveData(datalist,savepath):
  book=Workbook(write_only=True)
  ws=[]
  ws.append(book.create_sheet(title=savepath));
  ws[0].append(["电影详情链接","封面图片链接","影片中文名","影片外国名","评分",
                 "评价数","概况","导演","主演","年份","地区","类别"])
  for i in datalist:
    ws[0].append([i[0],i[1],i[2],i[3],i[4],i[5],i[6],i[7],i[8],i[9],i[10],i[11]])
  book.save(savepath)#保存

def main():
  baseurl='https://movie.douban.com/top250?start='
  datalist=getData(baseurl)
  savapath='豆瓣电影Top250.xlsx'
  saveData(datalist,savapath)

main()